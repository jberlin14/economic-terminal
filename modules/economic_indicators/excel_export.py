"""
Economic Indicators Excel Exporter

Exports indicator data to Excel with formatting and transformations.
"""

from io import BytesIO
from datetime import datetime, date
from typing import List, Optional, Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from loguru import logger

from .config import DASHBOARDS, get_all_indicators
from .storage import IndicatorStorage
from ..data_storage.database import get_db_context


class ExcelExporter:
    """
    Exports economic indicator data to formatted Excel files.
    """

    def __init__(self):
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.title_font = Font(size=14, bold=True)
        self.subtitle_font = Font(size=11, italic=True)

    def _format_worksheet(self, ws, title: str, subtitle: Optional[str] = None):
        """Apply formatting to worksheet headers"""
        # Title
        ws['A1'] = title
        ws['A1'].font = self.title_font

        if subtitle:
            ws['A2'] = subtitle
            ws['A2'].font = self.subtitle_font

    def _format_header_row(self, ws, row_num: int, columns: int):
        """Format header row with colors"""
        for col in range(1, columns + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')

    def _auto_size_columns(self, ws):
        """Auto-size all columns based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def export_single_series(
        self,
        series_id: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        include_transformations: List[str] = None
    ) -> bytes:
        """
        Export a single series to Excel with optional transformations.

        Args:
            series_id: FRED series ID
            start_date: Start date filter
            end_date: End date filter
            include_transformations: List of transforms to include (e.g., ['mom_percent', 'yoy_percent', 'ma_3'])

        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active

        with get_db_context() as db:
            storage = IndicatorStorage(db)

            # Get indicator metadata
            indicator = storage.get_indicator(series_id)
            if not indicator:
                raise ValueError(f"Series {series_id} not found")

            # Get data
            if include_transformations:
                df = storage.get_values_with_transforms(
                    series_id, start_date, end_date, include_transformations
                )
            else:
                df = storage.get_values(series_id, start_date, end_date)

            if df.empty:
                raise ValueError(f"No data available for {series_id}")

            # Format title
            title = f"{indicator.name} ({series_id})"
            subtitle = f"Units: {indicator.units} | Frequency: {indicator.frequency} | Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            self._format_worksheet(ws, title, subtitle)

            # Write data starting at row 4
            start_row = 4
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    # Format dates
                    if c_idx == 1 and r_idx > start_row:  # Date column, not header
                        if isinstance(value, date):
                            cell.number_format = 'YYYY-MM-DD'
                    # Format numbers
                    elif c_idx > 1 and r_idx > start_row:  # Value columns, not header
                        if isinstance(value, (int, float)):
                            cell.number_format = '0.0000'

            # Format header row
            self._format_header_row(ws, start_row, len(df.columns))

            # Auto-size columns
            self._auto_size_columns(ws)

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def export_multiple_series(
        self,
        series_ids: List[str],
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        format: str = 'separate_sheets'
    ) -> bytes:
        """
        Export multiple series to Excel.

        Args:
            series_ids: List of FRED series IDs
            start_date: Start date filter
            end_date: End date filter
            format: 'separate_sheets' or 'columns'

        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        with get_db_context() as db:
            storage = IndicatorStorage(db)

            if format == 'separate_sheets':
                # Each series gets its own sheet
                for series_id in series_ids:
                    indicator = storage.get_indicator(series_id)
                    if not indicator:
                        logger.warning(f"Series {series_id} not found, skipping")
                        continue

                    df = storage.get_values(series_id, start_date, end_date)
                    if df.empty:
                        logger.warning(f"No data for {series_id}, skipping")
                        continue

                    # Create sheet
                    ws = wb.create_sheet(title=series_id[:31])  # Excel sheet name limit

                    # Title
                    title = f"{indicator.name}"
                    subtitle = f"{series_id} | {indicator.units}"
                    self._format_worksheet(ws, title, subtitle)

                    # Write data
                    start_row = 4
                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
                        for c_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=r_idx, column=c_idx, value=value)
                            if c_idx == 1 and r_idx > start_row:
                                if isinstance(value, date):
                                    cell.number_format = 'YYYY-MM-DD'
                            elif c_idx > 1 and r_idx > start_row:
                                if isinstance(value, (int, float)):
                                    cell.number_format = '0.0000'

                    self._format_header_row(ws, start_row, 2)
                    self._auto_size_columns(ws)

            else:  # format == 'columns'
                # All series in one sheet with date index
                df_comparison = storage.get_comparison_data(series_ids, start_date, end_date)
                if df_comparison.empty:
                    raise ValueError("No data available for any series")

                ws = wb.create_sheet(title="Comparison")

                # Title
                title = "Economic Indicators Comparison"
                subtitle = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                self._format_worksheet(ws, title, subtitle)

                # Reset index to make date a column
                df_comparison = df_comparison.reset_index()
                df_comparison.columns = ['Date'] + [f"{sid}" for sid in df_comparison.columns[1:]]

                # Write data
                start_row = 4
                for r_idx, row in enumerate(dataframe_to_rows(df_comparison, index=False, header=True), start_row):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        if c_idx == 1 and r_idx > start_row:
                            if isinstance(value, date):
                                cell.number_format = 'YYYY-MM-DD'
                        elif c_idx > 1 and r_idx > start_row:
                            if isinstance(value, (int, float)):
                                cell.number_format = '0.0000'

                self._format_header_row(ws, start_row, len(df_comparison.columns))
                self._auto_size_columns(ws)

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def export_report_group(
        self,
        report_group: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> bytes:
        """
        Export all indicators in a report group.

        Args:
            report_group: Report group name (e.g., 'CPI Report')
            start_date: Start date filter
            end_date: End date filter

        Returns:
            Excel file as bytes
        """
        with get_db_context() as db:
            storage = IndicatorStorage(db)
            indicators = storage.get_indicators_by_report(report_group)

            if not indicators:
                raise ValueError(f"No indicators found for report group: {report_group}")

            series_ids = [ind.series_id for ind in indicators]
            return self.export_multiple_series(series_ids, start_date, end_date, format='separate_sheets')

    def export_dashboard(
        self,
        dashboard_name: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> bytes:
        """
        Export a pre-configured dashboard.

        Args:
            dashboard_name: Dashboard name ('inflation', 'labor', 'claims', 'gdp')
            start_date: Start date filter
            end_date: End date filter

        Returns:
            Excel file as bytes
        """
        if dashboard_name not in DASHBOARDS:
            raise ValueError(f"Dashboard {dashboard_name} not found")

        dashboard_config = DASHBOARDS[dashboard_name]
        series_ids = dashboard_config['series']
        default_transform = dashboard_config.get('default_transform', 'raw')

        wb = Workbook()
        wb.remove(wb.active)

        with get_db_context() as db:
            storage = IndicatorStorage(db)

            # Create summary sheet
            summary_ws = wb.create_sheet(title="Summary")
            summary_ws['A1'] = dashboard_config['name']
            summary_ws['A1'].font = self.title_font
            summary_ws['A2'] = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            summary_ws['A2'].font = self.subtitle_font

            # List all series
            summary_ws['A4'] = "Included Series:"
            summary_ws['A4'].font = Font(bold=True)
            for idx, series_id in enumerate(series_ids, 5):
                indicator = storage.get_indicator(series_id)
                if indicator:
                    summary_ws[f'A{idx}'] = f"{series_id}: {indicator.name}"

            # Create sheet for each series
            for series_id in series_ids:
                indicator = storage.get_indicator(series_id)
                if not indicator:
                    continue

                # Get data with default transform if applicable
                if default_transform != 'raw':
                    df = storage.get_values_with_transforms(
                        series_id, start_date, end_date, [default_transform]
                    )
                else:
                    df = storage.get_values(series_id, start_date, end_date)

                if df.empty:
                    continue

                # Create sheet
                ws = wb.create_sheet(title=series_id[:31])

                # Title
                title = f"{indicator.name}"
                subtitle = f"{series_id} | Transform: {default_transform}"
                self._format_worksheet(ws, title, subtitle)

                # Write data
                start_row = 4
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        if c_idx == 1 and r_idx > start_row:
                            if isinstance(value, date):
                                cell.number_format = 'YYYY-MM-DD'
                        elif c_idx > 1 and r_idx > start_row:
                            if isinstance(value, (int, float)):
                                cell.number_format = '0.0000'

                self._format_header_row(ws, start_row, len(df.columns))
                self._auto_size_columns(ws)

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def create_metadata_sheet(self, wb: Workbook, series_ids: List[str]):
        """Create a metadata summary sheet"""
        ws = wb.create_sheet(title="Metadata", index=0)

        with get_db_context() as db:
            storage = IndicatorStorage(db)

            # Title
            ws['A1'] = "Export Metadata"
            ws['A1'].font = self.title_font
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = self.subtitle_font

            # Headers
            headers = ['Series ID', 'Name', 'Units', 'Frequency', 'Latest Value', 'Latest Date']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.fill = self.header_fill
                cell.font = self.header_font

            # Data
            for idx, series_id in enumerate(series_ids, 5):
                indicator = storage.get_indicator(series_id)
                if indicator:
                    ws.cell(row=idx, column=1, value=indicator.series_id)
                    ws.cell(row=idx, column=2, value=indicator.name)
                    ws.cell(row=idx, column=3, value=indicator.units)
                    ws.cell(row=idx, column=4, value=indicator.frequency)
                    ws.cell(row=idx, column=5, value=indicator.latest_value)
                    if indicator.latest_date:
                        ws.cell(row=idx, column=6, value=indicator.latest_date.isoformat())

            self._auto_size_columns(ws)
